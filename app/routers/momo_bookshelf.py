"""
모모의 책장 - 분기별/학년별 주차 커리큘럼
GET  /momo-bookshelf                 -> 목록 조회 (연도·분기·학년 필터)
GET  /momo-bookshelf/upload           -> 업로드 화면
POST /momo-bookshelf/upload           -> 엑셀 업로드 -> "연간 전체 리스트" 시트 파싱 -> upsert
GET  /momo-bookshelf/required         -> 필독서 목록 (주차 뺀 학년+분기별 고유 도서 + ISBN)
POST /momo-bookshelf/required/sync    -> 커리큘럼에서 필독서 목록 재동기화
POST /momo-bookshelf/required/auto-link -> ISBN 없는 필독서 일괄 자동 매칭(제목/저자 유사도 확인)

업로드 파일은 "모모의책장_DB_..._연간_통합_주차별.xlsx"와 같은 형식으로,
"연간 전체 리스트" 시트에 분기/학년/주차/기간/도서명/저자(역자)/출판사
7개 열이 있어야 함. 같은 (연도, 분기, 학년, 주차) 조합이 이미 있으면
덮어쓰기(upsert)되므로 파일을 다시 올려 수정하는 것도 안전함.
"""
import io
import re
import asyncio
import difflib
from datetime import datetime

import httpx
from fastapi import APIRouter, Depends, Request, Form, File, UploadFile, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import distinct
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.momo_bookshelf import MomoBookshelfWeek, MomoRequiredBook, GRADES

router = APIRouter(prefix="/momo-bookshelf")
templates = Jinja2Templates(directory="app/templates")

SHEET_NAME = "연간 전체 리스트"
REQUIRED_HEADERS = ["분기", "학년", "주차", "기간", "도서명", "저자(역자)", "출판사"]

AUTO_LINK_BATCH_SIZE = 40    # 타임아웃 방지를 위해 한 번 호출에 처리할 최대 건수
AUTO_LINK_CONCURRENCY = 3    # 너무 높으면 알라딘이 일시적으로 빈 응답을 줄 수 있음(요청 폭주)
AUTO_LINK_REQUEST_DELAY = 0.2  # 각 요청 사이 최소 간격(초), 순간 폭주 방지


def _parse_week_number(raw) -> int | None:
    if raw is None:
        return None
    m = re.search(r"\d+", str(raw))
    return int(m.group()) if m else None


def _strip_extension_marker(title: str) -> str:
    """"이솝 이야기 (연장)" -> "이솝 이야기" (같은 책을 연속으로 읽는 재당 주차 표시 제거)"""
    return re.sub(r"\s*\(연장\)\s*$", "", title or "").strip()


_REQUIRED_TRAILING_NOISE_TOKENS = [
    r"[가-힣]{1,4}T",       # 담당 강사 코드 (예: 문요한T, 문T)
    r"교사용|학생용",
    r"수업\s*준비",
    r"\d+\s*(?:주차|차시)",
    # 단독 숫자(예: \d+)는 일부러 제외함: "한국 대표 단편문학선 1"/"2"처럼
    # 뒤에 붙은 숫자가 서로 다른 책(권)을 가리키는 경우가 있어 여기서 지우면 안 됨.
]


def _clean_required_title(title: str) -> str:
    """(연장) 표시 + "N주차_교사용"/강사코드 등 수업 메타데이터를 제거.
    /isbn 검색 프리필에 쓰는 _clean_material_title_for_search와 달리 단독 숫자는 지우지 않음
    (권 번호가 붙은 여러 권짜리 도서를 서로 다른 책으로 유지하기 위함)."""
    from app.isbn import _LEADING_TITLE_NOISE_RE

    t = _strip_extension_marker(title)
    t = _LEADING_TITLE_NOISE_RE.sub("", t, count=1)

    changed = True
    while changed:
        changed = False
        for token in _REQUIRED_TRAILING_NOISE_TOKENS:
            m = re.search(rf"[\s_]*(?:{token})[\s_]*$", t)
            if m and m.start() > 0:
                t = t[:m.start()]
                changed = True
                break

    cleaned = t.strip(" _'\"‘’“”")
    return cleaned if cleaned else (title or "").strip()


def _sync_required_books(db: Session) -> dict:
    """momo_bookshelf_weeks에서 주차를 빼고 학년+분기별 고유 도서만 뽑아 momo_required_books에 반영.
    "(연장)"/"N주차_교사용" 등 재당 주차 표시는 원본과 합치고, 휴강/특강 표시 행은 제외함.
    같은 학년+분기 안에서 제목이 "..."로 끝나면(엑셀에 "위와 동일" 뜻으로 흔히 씀) 바로 이전 주차의
    책 제목을 이어받은 것으로 보고 합침.
    이미 연결된 ISBN 등 기존 필드는 건드리지 않음. 예전 방식으로 잘못 쪼개져 남아있던 중복 행은
    새 제목 기준으로 정리하되, ISBN이 연결돼 있으면 잃지 않고 정본 행으로 옮겨줌."""
    weeks = (
        db.query(MomoBookshelfWeek)
        .filter(MomoBookshelfWeek.is_holiday == False)
        .order_by(MomoBookshelfWeek.year, MomoBookshelfWeek.quarter, MomoBookshelfWeek.grade, MomoBookshelfWeek.week_number)
        .all()
    )

    unique = {}  # (year, quarter, grade, clean_title) -> (author, publisher)
    last_title_by_group = {}  # (year, quarter, grade) -> 그 그룹에서 마지막으로 확정된 clean_title
    for w in weeks:
        raw_clean = _clean_required_title(w.title)
        if not raw_clean:
            continue
        group_key = (w.year, w.quarter, w.grade)

        clean_title = raw_clean
        if raw_clean.endswith("...") and group_key in last_title_by_group:
            prefix = raw_clean[:-3].strip()
            prev_title = last_title_by_group[group_key]
            if not prefix or prev_title.startswith(prefix):
                clean_title = prev_title  # "..."는 이전 주차와 같은 책의 연속으로 봄

        last_title_by_group[group_key] = clean_title
        key = (w.year, w.quarter, w.grade, clean_title)
        if key not in unique:
            unique[key] = (w.author, w.publisher)

    added, updated = 0, 0
    for (year, quarter, grade, title), (author, publisher) in unique.items():
        existing = db.query(MomoRequiredBook).filter(
            MomoRequiredBook.year == year,
            MomoRequiredBook.quarter == quarter,
            MomoRequiredBook.grade == grade,
            MomoRequiredBook.title == title,
        ).first()
        if existing:
            if existing.author != author or existing.publisher != publisher:
                existing.author = author
                existing.publisher = publisher
                existing.updated_at = datetime.now()
                updated += 1
        else:
            db.add(MomoRequiredBook(year=year, quarter=quarter, grade=grade, title=title, author=author, publisher=publisher))
            added += 1
    db.commit()

    merged, removed = _reconcile_orphan_required_books(db, unique)

    return {"added": added, "updated": updated, "total": len(unique), "merged": merged, "removed": removed}


def _reconcile_orphan_required_books(db: Session, unique: dict) -> tuple[int, int]:
    """새 제목 기준(unique)에 더 이상 없는 기존 필독서 행(예전 분리 방식의 잔재)을 정리.
    ISBN이 연결돼 있으면 같은 학년+분기의 정본 행으로 옮긴 뒤(정본에 ISBN이 없을 때만) 삭제하고,
    ISBN이 없으면 그냥 삭제함."""
    valid_keys = set(unique.keys())
    all_rows = db.query(MomoRequiredBook).all()
    orphans = [r for r in all_rows if (r.year, r.quarter, r.grade, r.title) not in valid_keys]
    if not orphans:
        return 0, 0

    canonical_by_group = {}
    for r in all_rows:
        canonical_by_group.setdefault((r.year, r.quarter, r.grade), []).append(r)

    merged = removed = 0
    for orphan in orphans:
        group = canonical_by_group.get((orphan.year, orphan.quarter, orphan.grade), [])
        target = next((c for c in group if c.id != orphan.id and c.title in orphan.title), None)
        if target and (orphan.isbn13 or orphan.isbn10) and not (target.isbn13 or target.isbn10):
            target.isbn13 = orphan.isbn13
            target.isbn10 = orphan.isbn10
            target.publisher = target.publisher or orphan.publisher
            target.cover_url = orphan.cover_url
            target.aladin_link = orphan.aladin_link
            target.is_auto_linked = orphan.is_auto_linked
            merged += 1
        db.delete(orphan)
        removed += 1

    db.commit()
    return merged, removed


def _parse_workbook(file_bytes: bytes) -> list[dict]:
    """엑셀 파일에서 '연간 전체 리스트' 시트를 읽어 행 딕셔너리 리스트로 반환. 형식이 안 맞으면 ValueError."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'"{SHEET_NAME}" 시트를 찾을 수 없습니다. (시트 목록: {", ".join(wb.sheetnames)})')

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    header_norm = [str(h).strip() if h else "" for h in (header or [])][:7]
    if header_norm != REQUIRED_HEADERS:
        raise ValueError(
            f"헤더가 예상과 다릅니다. 필요한 열: {', '.join(REQUIRED_HEADERS)} / 실제: {', '.join(header_norm)}"
        )

    parsed = []
    for row in rows_iter:
        if not row or not any(row):
            continue
        padded = (tuple(row) + (None,) * 7)[:7]
        quarter, grade, week_raw, date_range, title, author, publisher = padded
        if not (quarter and grade and title):
            continue
        week_number = _parse_week_number(week_raw)
        if week_number is None:
            continue
        author_s = str(author).strip() if author else None
        parsed.append({
            "quarter": str(quarter).strip(),
            "grade": str(grade).strip(),
            "week_number": week_number,
            "date_range": str(date_range).strip() if date_range else None,
            "title": str(title).strip(),
            "author": author_s,
            "publisher": str(publisher).strip() if publisher else None,
            "is_holiday": author_s == "-",
        })
    return parsed


@router.get("", response_class=HTMLResponse)
def bookshelf_list(
    request: Request,
    year: str | None = None,
    quarter: str | None = None,
    grade: str | None = None,
    db: Session = Depends(get_db),
):
    """연도/분기/학년으로 필터링한 커리큘럼 목록"""
    # "전체" 옵션 선택 시 빈 문자열로 넘어오므로 int 파싱 전에 방어
    year_val = int(year) if year else None

    query = db.query(MomoBookshelfWeek)
    if year_val:
        query = query.filter(MomoBookshelfWeek.year == year_val)
    if quarter:
        query = query.filter(MomoBookshelfWeek.quarter == quarter)
    if grade:
        query = query.filter(MomoBookshelfWeek.grade == grade)

    weeks = query.order_by(
        MomoBookshelfWeek.grade,
        MomoBookshelfWeek.quarter,
        MomoBookshelfWeek.week_number,
    ).all()

    years = [r[0] for r in db.query(distinct(MomoBookshelfWeek.year)).order_by(MomoBookshelfWeek.year.desc()).all()]
    quarters = sorted(r[0] for r in db.query(distinct(MomoBookshelfWeek.quarter)).all())

    return templates.TemplateResponse("momo_bookshelf/list.html", {
        "request": request,
        "weeks": weeks,
        "years": years,
        "quarters": quarters,
        "grades": GRADES,
        "filter_year": year_val,
        "filter_quarter": quarter,
        "filter_grade": grade,
        "total_count": db.query(MomoBookshelfWeek).count(),
    })


@router.get("/upload", response_class=HTMLResponse)
def bookshelf_upload_page(request: Request):
    return templates.TemplateResponse("momo_bookshelf/upload.html", {
        "request": request,
        "default_year": datetime.now().year,
    })


@router.post("/upload")
async def bookshelf_upload(
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
    year: int = Form(...),
):
    """같은 양식의 엑셀 파일을 업로드해 커리큘럼을 등록/갱신 (같은 연도+분기+학년+주차는 덮어씀)"""
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail=".xlsx 파일만 업로드할 수 있습니다.")

    content = await file.read()
    try:
        rows = _parse_workbook(content)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not rows:
        raise HTTPException(status_code=400, detail="파싱된 데이터가 없습니다. 파일 형식을 확인하세요.")

    added, updated = 0, 0
    for r in rows:
        existing = db.query(MomoBookshelfWeek).filter(
            MomoBookshelfWeek.year == year,
            MomoBookshelfWeek.quarter == r["quarter"],
            MomoBookshelfWeek.grade == r["grade"],
            MomoBookshelfWeek.week_number == r["week_number"],
        ).first()
        if existing:
            existing.date_range = r["date_range"]
            existing.title = r["title"]
            existing.author = r["author"]
            existing.publisher = r["publisher"]
            existing.is_holiday = r["is_holiday"]
            existing.updated_at = datetime.now()
            updated += 1
        else:
            db.add(MomoBookshelfWeek(year=year, **r))
            added += 1

    db.commit()
    sync_result = _sync_required_books(db)
    return JSONResponse({
        "ok": True, "added": added, "updated": updated, "total": len(rows),
        "required_books": sync_result,
    })


@router.get("/required", response_class=HTMLResponse)
def required_books_list(
    request: Request,
    year: str | None = None,
    quarter: str | None = None,
    grade: str | None = None,
    db: Session = Depends(get_db),
):
    """모모의 책장 필독서 목록 (커리큘럼에서 주차를 뺀 학년+분기별 고유 도서 + ISBN)"""
    year_val = int(year) if year else None

    query = db.query(MomoRequiredBook)
    if year_val:
        query = query.filter(MomoRequiredBook.year == year_val)
    if quarter:
        query = query.filter(MomoRequiredBook.quarter == quarter)
    if grade:
        query = query.filter(MomoRequiredBook.grade == grade)

    books = query.order_by(
        MomoRequiredBook.grade, MomoRequiredBook.quarter, MomoRequiredBook.title
    ).all()

    years = [r[0] for r in db.query(distinct(MomoRequiredBook.year)).order_by(MomoRequiredBook.year.desc()).all()]
    quarters = sorted(r[0] for r in db.query(distinct(MomoRequiredBook.quarter)).all())

    return templates.TemplateResponse("momo_bookshelf/required.html", {
        "request": request,
        "books": books,
        "years": years,
        "quarters": quarters,
        "grades": GRADES,
        "filter_year": year_val,
        "filter_quarter": quarter,
        "filter_grade": grade,
        "total_count": db.query(MomoRequiredBook).count(),
        "linked_count": db.query(MomoRequiredBook).filter(
            (MomoRequiredBook.isbn13.isnot(None)) | (MomoRequiredBook.isbn10.isnot(None))
        ).count(),
    })


@router.post("/required/sync")
def required_books_sync(db: Session = Depends(get_db)):
    """커리큘럼(momo_bookshelf_weeks)에서 필독서 목록을 다시 동기화 (수동 재실행용)"""
    result = _sync_required_books(db)
    return JSONResponse({"ok": True, **result})


def _normalize_for_match(s: str) -> str:
    return re.sub(r"[\s\W_]+", "", s or "").lower()


def _is_confident_match(req_title: str, req_author: str | None, cand_title: str, cand_author: str) -> bool:
    """제목/저자가 충분히 비슷하면 같은 책으로 판단 (필독서 제목은 커리큘럼에서 이미 정제된 값이라
    독서논술 교재 제목보다 훨씬 깨끗해서 이 정도 기준으로도 오매칭 위험이 낮음)"""
    nt1, nt2 = _normalize_for_match(req_title), _normalize_for_match(cand_title)
    if not nt1 or not nt2:
        return False

    title_match = nt1 == nt2 or nt1 in nt2 or nt2 in nt1
    if not title_match:
        title_match = difflib.SequenceMatcher(None, nt1, nt2).ratio() >= 0.8
    if not title_match:
        return False

    # 저자 정보가 있는데 서로 완전히 다르고 제목도 정확히 일치하지 않으면 동명이서로 보고 보류
    if req_author:
        na1, na2 = _normalize_for_match(req_author), _normalize_for_match(cand_author)
        if na1 and na2 and na1 not in na2 and na2 not in na1 and nt1 != nt2:
            return False
    return True


@router.post("/required/auto-link")
async def required_books_auto_link(db: Session = Depends(get_db)):
    """
    ISBN이 없는 필독서를 대상으로 알라딘 검색 1위 결과가 제목/저자와 충분히 비슷하면 자동 연결.
    애매하면 건너뛰고 그대로 두어 /isbn 화면에서 수동 확인하도록 함.
    타임아웃 방지를 위해 한 번 호출에 최대 AUTO_LINK_BATCH_SIZE건만 처리 (남으면 다시 눌러서 이어서 처리).
    """
    from app.isbn import _search_aladin

    targets = (
        db.query(MomoRequiredBook)
        .filter(MomoRequiredBook.isbn13.is_(None), MomoRequiredBook.isbn10.is_(None))
        .order_by(MomoRequiredBook.created_at)
        .limit(AUTO_LINK_BATCH_SIZE)
        .all()
    )
    if not targets:
        return JSONResponse({"ok": True, "linked": 0, "skipped": 0, "processed": 0, "remaining": 0})

    semaphore = asyncio.Semaphore(AUTO_LINK_CONCURRENCY)

    async def _find_match(client: httpx.AsyncClient, book: MomoRequiredBook):
        async with semaphore:
            await asyncio.sleep(AUTO_LINK_REQUEST_DELAY)
            try:
                # 제목만으로 검색 (제목+저자를 합쳐서 검색하면 저자 표기가 조금만 달라도
                # 결과가 아예 0건이 되는 경우가 많음). 저자는 매칭 단계에서만 확인.
                res = await _search_aladin(client, book.title, None, 5)
            except HTTPException:
                return (book.id, None)
            for cand in res["results"]:
                if _is_confident_match(book.title, book.author, cand["title"], cand["author"]):
                    return (book.id, cand)
            return (book.id, None)

    async with httpx.AsyncClient(timeout=10) as client:
        matches = await asyncio.gather(*[_find_match(client, b) for b in targets])

    linked = 0
    for book_id, cand in matches:
        if cand is None:
            continue
        # 아주 오래된 책은 알라딘에도 isbn13/isbn10이 둘 다 없을 수 있음 - 그런 경우는
        # "매칭"이어도 저장할 ISBN이 없으므로 연결로 치지 않음(안 그러면 다음 라운드에도
        # 계속 미연결로 잡혀 똑같은 시도를 무한 반복하게 됨)
        if not (cand.get("isbn13") or cand.get("isbn10")):
            continue
        book = db.query(MomoRequiredBook).filter(MomoRequiredBook.id == book_id).first()
        if not book:
            continue
        book.isbn13 = cand.get("isbn13") or None
        book.isbn10 = cand.get("isbn10") or None
        book.publisher = book.publisher or cand.get("publisher") or None
        book.cover_url = cand.get("cover") or None
        book.aladin_link = cand.get("link") or None
        book.is_auto_linked = True
        book.updated_at = datetime.now()
        linked += 1

    db.commit()
    remaining = db.query(MomoRequiredBook).filter(
        MomoRequiredBook.isbn13.is_(None), MomoRequiredBook.isbn10.is_(None)
    ).count()
    return JSONResponse({
        "ok": True, "linked": linked, "skipped": len(targets) - linked,
        "processed": len(targets), "remaining": remaining,
    })
