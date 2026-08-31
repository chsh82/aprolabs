"""raw/schema-reading/ 엑셀 2개(학습 도구어 사전, 스키마 어휘 목록) 파싱.

docs/literacy/04-스키마리딩어휘적재.md 4절 참고. 어휘퀴즈DB.xlsx(3번째
파일, 4지선다 338문항)는 이 모듈에서 다루지 않는다 - quiz_items 대상이라
별도 지시서로 처리한다(docs/literacy/04-스키마리딩어휘.md "미적재 자료" 참고).

실제 파일을 열어서 확인한 결과 문서 설명과 다른 부분이 있어 그대로 반영했다:
- 학습 도구어 사전의 헤더는 3행째가 아니라 **4행째**다(3행째는 'level N'
  라벨 텍스트일 뿐). 4행을 헤더로 읽으면 문서가 경고한 "첫 행이 헤더로
  중복되는 문제" 자체가 발생하지 않는다.
- (유)/(반) 태그는 Level3에만 있다(116/146건). Level1·2·5·6은 태그가
  전혀 없어 유의어/반의어 방향을 알 수 없다 - 태그 없는 항목은
  'tooldict-related'(방향 불명)로 저장한다(사용자 결정).
- tooldict 표제어 뒤에 괄호가 붙은 9건 처리(사용자 결정, dry-run에서 발견):
  괄호 안이 순수 한자(CJK)면 한자 원어로 보고 `origin`으로 분리
  (`지(知)`, `내(內)`). 한자가 아니면 용법 예시로 보고 표제어에서 떼어
  `usage_note`로 보관(`긋다(밑줄을)`, `뽑다(카드를)`, `생활(방식)`,
  `세우다(식을)`) - 최종적으로 `note`에 "용법: ..."로 기록한다.
- **같은 레벨 내부에서도 표제어가 중복되는 경우가 42건 있고, 이 중 16건은
  단순 중복이 아니라 실제 동형이의어(뜻풀이가 서로 다름)임을 발견함**
  (사용자 결정). Phase 3의 "동레벨 중복은 조용히 병합" 규칙은 완전히
  같은 내용이 재등장하는 경우를 가정한 것이라 여기 안 맞는다. 그래서
  버리지 않고, 기존 `|`-다의어 처리와 같은 방식으로 대표 항목의
  definitions에 이어붙여 전부 `examples`로 보존한다
  (`merge_same_level_duplicates` 참고). 딱 1건("삶", L4)은 다른 단어의
  뜻풀이가 잘못 들어간 원본 데이터 오류로 판단해 보존하지 않고 제외한다
  (`_KNOWN_BAD_DEFINITIONS`) - 원본 엑셀은 수정하지 않는다.
"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).resolve().parents[2] / "raw" / "schema-reading"
TOOLDICT_PATH = RAW_DIR / "학습 도구어 사전 레벨1~6 (0925최종버전).xlsx"
SCHEMA_PATH = RAW_DIR / "스키마 어휘 목록(레벨2~5 완성).xlsx"

_TAG_RE = re.compile(r"^(.*)\((유|반)\)$")
_SHEET_KEY = {"사회 스키마": "social", "과학 스키마": "sci", "인문철학스키마": "phil"}

_TRAILING_PAREN_RE = re.compile(r"^(.+?)\(([^)]+)\)$")
_PURE_HANJA_RE = re.compile(r"^[一-鿿]+$")


def _split_trailing_paren(headword: str) -> tuple[str, str | None, str | None]:
    """표제어 뒤 괄호를 떼어낸다. 괄호 안이 순수 한자면 (표제어, 한자, None),
    아니면 용법 예시로 보고 (표제어, None, 용법텍스트)를 반환한다.
    괄호가 없으면 (원본 그대로, None, None)."""
    m = _TRAILING_PAREN_RE.match(headword)
    if not m:
        return headword, None, None
    base, paren_content = m.group(1).strip(), m.group(2).strip()
    if _PURE_HANJA_RE.match(paren_content):
        return base, paren_content, None
    return base, None, paren_content


@dataclass
class ToolDictEntry:
    level: int
    week: str | None
    headword: str
    raw_headword: str  # 정제(strip) 전 원본 - dry-run 비교용
    definitions: list[str]
    pos: str | None
    synonyms: list[str] = field(default_factory=list)
    antonyms: list[str] = field(default_factory=list)
    related: list[str] = field(default_factory=list)  # (유)/(반) 태그 없음 - 방향 불명
    row_no: int = 0
    origin: str | None = None       # 표제어 뒤 괄호가 순수 한자였던 경우
    usage_note: str | None = None   # 표제어 뒤 괄호가 용법 예시였던 경우
    data_error_note: str | None = None  # 같은 레벨 병합 중 제외한 원본 데이터 오류 안내


@dataclass
class SchemaEntry:
    sheet_key: str  # social/sci/phil
    row_no: int
    headword: str
    raw_headword: str  # 정제(strip) 전 원본 - dry-run 비교용
    definition: str | None
    sense_category: str | None  # 중분류, 없으면 대분류로 대체
    subject_category: str | None
    level: int | None
    sub_category: str | None  # note용(소분류)
    week: str | None  # note용(주차)


@dataclass
class SameLevelMerge:
    """같은 (level, headword) 안에서 실제로 다른 뜻풀이가 여럿 있어 병합한 결과 -
    dry-run 보고용(대표로 남은 뜻풀이 vs examples로 밀려난 뜻풀이를 나란히 보여줌)."""
    level: int
    headword: str
    primary_row_no: int
    primary_definitions: list[str]
    pushed_definitions: list[str]  # examples로 밀려난 것(대표 항목이 원래 갖고 있던 것 제외)
    excluded_error_definitions: list[str]  # 원본 데이터 오류로 제외된 것


# 병합하지 않고 제외할 것으로 판단한 원본 데이터 오류(사용자가 직접 지목한 건만
# 하드코딩 - 일반화된 오류 탐지 로직은 만들지 않는다). 원본 엑셀은 수정하지 않는다.
# 주의: 이 목록이 계속 늘어나면(현재 2건) 더 늘어날 경우 이 파일 밖(JSON/CSV 등)
# 으로 분리하는 게 낫다 - 지금 규모에서는 하드코딩으로 충분하다고 판단.
_KNOWN_BAD_DEFINITIONS: dict[tuple[int, str], set[str]] = {
    (4, "삶"): {"서로 관련을 가짐. 또는 그 관계."},
    (1, "정하다"): {
        "겉으로 보기에 비뚤어지거나 굽은 데가 없다.",
        "사리에 맞고 바르다.",
    },  # '곧다'의 뜻풀이가 잘못 들어감 - 사용자 확인
}

# 기본값은 row_no가 가장 작은 행을 대표로 삼지만, "행번호 순서 = 중요도 순서"가
# 아닌 경우가 있어(예: "긋다"는 row61이 "비가 그치다"라는 덜 중심적인 뜻) 사용자가
# 직접 확인하고 지정한 건만 여기 하드코딩한다.
_MANUAL_PRIMARY_OVERRIDES: dict[tuple[int, str], int] = {
    (1, "긋다"): 63,  # "금이나 줄을 그리다" - 사용자 결정
    (1, "무리"): 138,  # "사람·짐승·사물이 모여 뭉친 동아리" - 사용자 결정(초3~4 수준에 맞음)
}


def merge_same_level_duplicates(
    entries: list[ToolDictEntry],
) -> tuple[list[ToolDictEntry], list[SameLevelMerge]]:
    """같은 (level, headword)로 여러 행이 있으면 하나로 합친다.

    - 뜻풀이가 완전히 같으면 단순 중복 제거(정보 손실 없음).
    - 뜻풀이가 실제로 다르면(동형이의어) 버리지 않고 대표 항목(가장 낮은
      row_no)의 definitions에 이어붙인다 - 기존 `|`-다의어 처리와 같은
      경로로 examples에 보존된다.
    - `_KNOWN_BAD_DEFINITIONS`에 등록된 뜻풀이는 병합하지 않고 제외하며
      `data_error_note`에 남긴다.
    - synonyms/antonyms/related, origin/usage_note도 합친다(대표에 없으면
      다른 행에서라도 채움).

    (합쳐진 목록, dry-run 보고용 SameLevelMerge 목록)을 반환한다. 보고 목록은
    실제로 뜻풀이가 여러 개 합쳐진 경우에만 담긴다(단순 중복 제거는 안 담음).
    """
    groups: dict[tuple[int, str], list[ToolDictEntry]] = defaultdict(list)
    for e in entries:
        groups[(e.level, e.headword)].append(e)

    merged: list[ToolDictEntry] = []
    reports: list[SameLevelMerge] = []

    for (level, headword), group in groups.items():
        if len(group) == 1:
            merged.append(group[0])
            continue

        override_row_no = _MANUAL_PRIMARY_OVERRIDES.get((level, headword))
        if override_row_no is not None:
            ranked = sorted(group, key=lambda e: (e.row_no != override_row_no, e.row_no))
        else:
            ranked = sorted(group, key=lambda e: e.row_no)
        primary = ranked[0]
        original_primary_definitions = list(primary.definitions)

        excluded_texts = _KNOWN_BAD_DEFINITIONS.get((level, headword), set())
        excluded: list[str] = []
        pushed: list[str] = []

        combined_definitions = list(primary.definitions)
        combined_synonyms = list(primary.synonyms)
        combined_antonyms = list(primary.antonyms)
        combined_related = list(primary.related)

        for other in ranked[1:]:
            for d in other.definitions:
                if d in excluded_texts:
                    excluded.append(d)
                    continue
                if d not in combined_definitions:
                    combined_definitions.append(d)
                    pushed.append(d)
            for s in other.synonyms:
                if s not in combined_synonyms:
                    combined_synonyms.append(s)
            for a in other.antonyms:
                if a not in combined_antonyms:
                    combined_antonyms.append(a)
            for r in other.related:
                if r not in combined_related:
                    combined_related.append(r)
            if primary.usage_note is None and other.usage_note:
                primary.usage_note = other.usage_note
            if primary.origin is None and other.origin:
                primary.origin = other.origin

        primary.definitions = combined_definitions
        primary.synonyms = combined_synonyms
        primary.antonyms = combined_antonyms
        primary.related = combined_related
        if excluded:
            primary.data_error_note = f"원본 데이터 오류로 제외된 뜻풀이 {len(excluded)}건 있음"

        if pushed or excluded:
            reports.append(SameLevelMerge(
                level=level,
                headword=headword,
                primary_row_no=primary.row_no,
                primary_definitions=original_primary_definitions,
                pushed_definitions=pushed,
                excluded_error_definitions=excluded,
            ))

        merged.append(primary)

    return merged, reports


def _split_multi(value) -> list[str]:
    """'|' 또는 개행으로 구분된 여러 값을 리스트로. 둘 다 구분자로 쓰인 셀이 있어(예:
    '과업(유)\\n숙제(유)') 둘 다 처리한다."""
    if value is None:
        return []
    parts = re.split(r"[|\n]", str(value))
    return [p.strip() for p in parts if p.strip()]


def _classify_related(items: list[str]) -> tuple[list[str], list[str], list[str]]:
    """(유)/(반) 태그로 synonyms/antonyms를 뽑고, 태그 없는 건 related(방향 불명)로."""
    synonyms, antonyms, related = [], [], []
    for item in items:
        m = _TAG_RE.match(item)
        if m:
            word = m.group(1).strip()
            (synonyms if m.group(2) == "유" else antonyms).append(word)
        elif item:
            related.append(item)
    return synonyms, antonyms, related


def _strip_tag(word: str) -> str:
    """분리된 유의어/반의어 컬럼(Level4)에도 태그가 섞여 있을 경우 텍스트만 남긴다(안전장치)."""
    m = _TAG_RE.match(word)
    return m.group(1).strip() if m else word


def iter_tooldict_entries():
    wb = openpyxl.load_workbook(TOOLDICT_PATH, data_only=True)
    for sheet_name in wb.sheetnames:
        level = int(sheet_name.replace("Level", ""))
        ws = wb[sheet_name]
        has_split_columns = ws.max_column >= 7  # Level4만 유의어/반의어 컬럼이 분리됨

        for row_no in range(5, ws.max_row + 1):  # 4행이 헤더, 5행부터 데이터
            headword = ws.cell(row=row_no, column=3).value
            if not headword or not str(headword).strip():
                continue

            week = ws.cell(row=row_no, column=2).value
            definitions = _split_multi(ws.cell(row=row_no, column=4).value)
            pos_values = _split_multi(ws.cell(row=row_no, column=5).value)
            pos = pos_values[0] if pos_values else None

            if has_split_columns:
                synonyms = [_strip_tag(s) for s in _split_multi(ws.cell(row=row_no, column=6).value)]
                antonyms = [_strip_tag(a) for a in _split_multi(ws.cell(row=row_no, column=7).value)]
                related: list[str] = []
            else:
                combined = _split_multi(ws.cell(row=row_no, column=6).value)
                synonyms, antonyms, related = _classify_related(combined)

            stripped = str(headword).strip()
            cleaned_headword, origin, usage_note = _split_trailing_paren(stripped)

            yield ToolDictEntry(
                level=level,
                week=str(week).strip() if week else None,
                headword=cleaned_headword,
                raw_headword=str(headword),
                definitions=definitions,
                pos=pos,
                synonyms=synonyms,
                antonyms=antonyms,
                related=related,
                row_no=row_no,
                origin=origin,
                usage_note=usage_note,
            )


def iter_schema_entries():
    wb = openpyxl.load_workbook(SCHEMA_PATH, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet_key = _SHEET_KEY[sheet_name]
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):  # 1행이 헤더, 2행부터 데이터
            headword = ws.cell(row=r, column=2).value
            if not headword or not str(headword).strip():
                continue

            row_no = ws.cell(row=r, column=1).value
            stage_raw = ws.cell(row=r, column=3).value
            week = ws.cell(row=r, column=4).value
            subject_category = ws.cell(row=r, column=5).value
            mid_category = ws.cell(row=r, column=6).value
            sub_category = ws.cell(row=r, column=7).value
            definition = ws.cell(row=r, column=8).value

            level_match = re.match(r"(\d+)", str(stage_raw)) if stage_raw else None

            yield SchemaEntry(
                sheet_key=sheet_key,
                row_no=int(row_no),
                headword=str(headword).strip(),
                raw_headword=str(headword),
                definition=(str(definition).strip() if definition and str(definition).strip() else None),
                sense_category=(
                    str(mid_category).strip() if mid_category and str(mid_category).strip()
                    else (str(subject_category).strip() if subject_category else None)
                ),
                subject_category=str(subject_category).strip() if subject_category else None,
                level=int(level_match.group(1)) if level_match else None,
                sub_category=str(sub_category).strip() if sub_category else None,
                week=str(week).strip() if week else None,
            )
